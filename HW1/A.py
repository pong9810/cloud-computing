from flask import Flask, jsonify, request
from flask_cors import CORS
import openpyxl
import os

app = Flask(__name__)
CORS(app)

EXCEL_FILE = 'salary.xlsx'

COLUMNS = [
    '員工編號', '姓名', '部門',
    '補習班時數', '補習班時薪',
    '家教時數', '家教時薪',
    '出差天數', '出差津貼',
    '獎金', '扣款'
]

def get_workbook():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(COLUMNS)
        wb.save(EXCEL_FILE)
    return openpyxl.load_workbook(EXCEL_FILE)

def rows_to_list(ws):
    employees = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            emp = {}
            for i, col in enumerate(COLUMNS):
                emp[col] = row[i] if row[i] is not None else 0
            employees.append(emp)
    return employees

# 取得全部員工
@app.route('/api/employees', methods=['GET'])
def get_employees():
    wb = get_workbook()
    ws = wb.active
    return jsonify(rows_to_list(ws))

# 取得單一員工
@app.route('/api/employees/<string:emp_id>', methods=['GET'])
def get_employee(emp_id):
    wb = get_workbook()
    ws = wb.active
    for emp in rows_to_list(ws):
        if emp['員工編號'] == emp_id:
            return jsonify(emp)
    return jsonify({'error': '找不到員工'}), 404

# 新增員工
@app.route('/api/employees', methods=['POST'])
def add_employee():
    data = request.json
    wb = get_workbook()
    ws = wb.active

    # 檢查編號是否重複
    for emp in rows_to_list(ws):
        if emp['員工編號'] == data.get('員工編號'):
            return jsonify({'error': '員工編號已存在'}), 400

    row = [data.get(col, 0) for col in COLUMNS]
    ws.append(row)
    wb.save(EXCEL_FILE)
    return jsonify({'message': f"員工 {data.get('姓名')} 新增成功", 'employee': data}), 201

# 修改員工
@app.route('/api/employees/<string:emp_id>', methods=['PUT'])
def update_employee(emp_id):
    data = request.json
    wb = get_workbook()
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        if row[0].value == emp_id:
            for i, col in enumerate(COLUMNS):
                if col in data:
                    row[i].value = data[col]
            wb.save(EXCEL_FILE)
            return jsonify({'message': f'{emp_id} 更新成功'})

    return jsonify({'error': '找不到員工'}), 404

# 刪除員工
@app.route('/api/employees/<string:emp_id>', methods=['DELETE'])
def delete_employee(emp_id):
    wb = get_workbook()
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        if row[0].value == emp_id:
            ws.delete_rows(row[0].row)
            wb.save(EXCEL_FILE)
            return jsonify({'message': f'{emp_id} 刪除成功'})

    return jsonify({'error': '找不到員工'}), 404

# 查詢員工（依姓名或編號）
@app.route('/api/employees/search', methods=['GET'])
def search_employees():
    keyword = request.args.get('keyword', '').strip()
    wb = get_workbook()
    ws = wb.active
    all_employees = rows_to_list(ws)
    result = [e for e in all_employees if keyword in e['員工編號'] or keyword in e['姓名']]
    return jsonify({'employees': result})

# 上傳 Excel
@app.route('/api/upload', methods=['POST'])
def upload_excel():
    if 'file' not in request.files:
        return jsonify({'error': '沒有收到檔案'}), 400

    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': '請上傳 Excel 檔案'}), 400

    file.save(EXCEL_FILE)
    return jsonify({'message': 'Excel 上傳成功'})

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)